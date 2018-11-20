import feedparser
import requests
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, NamedStyle

url_for_parse = 'http://zakupki.gov.ru/epz/order/extendedsearch/rss?searchString=%D0%BE%D1%82%D0%B1%D0%BE%D1%80&morphology=on&openMode=USE_DEFAULT_PARAMS&pageNumber=1&sortDirection=false&recordsPerPage=_10&showLotsInfoHidden=false&ppRf615=on&af=on&currencyIdGeneral=-1&regionDeleted=false&oktmoIdsWithNested=on&sortBy=UPDATE_DATE'
user_agent_string = 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'

# Создаем екзель таблицу
wb = Workbook()
ws = wb.active

#Стиль заголовка
styleheader = NamedStyle(name='styleheader')
styleheader.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
styleheader.border = Border (  left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))
wb.add_named_style(styleheader)
#Стиль остальных строк
styleraws = NamedStyle(name='styleraws')
styleraws.alignment = Alignment(wrap_text=True, shrink_to_fit=True)
styleraws.border = Border (  left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))
wb.add_named_style(styleraws)


ws['A1']= '№пп'
ws.column_dimensions['A'].width= 4
ws['B1']= 'Название отбора'
ws.column_dimensions['B'].width= 34.14
ws['C1']= 'Ссылка'
ws.column_dimensions['C'].width= 36.71
ws['D1']= 'Тема отбора'
ws.column_dimensions['D'].width= 66.71
ws['E1']= 'Организатор'
ws.column_dimensions['E'].width= 24.71
ws['F1']= 'Регион'
ws.column_dimensions['F'].width= 15.29
ws['G1']= 'Начало подачи'
ws.column_dimensions['G'].width= 10.43
ws['H1']= 'Конец подачи'
ws.column_dimensions['H'].width= 10.43
for col in ws['A1:H1']:
    for cell in col:
        cell.style = 'styleheader'

d = feedparser.parse(url_for_parse)
number_of_rec: int = 0
for x in d['entries']:
    number_of_rec += 1
    print(str(number_of_rec) + ' ------------------------------------------')
    ws['A' + str(number_of_rec + 1)] = str(number_of_rec)
    print(x['title'])
    ws['B' + str(number_of_rec + 1)] = x['title']
    print(x['link'])
    ws['C' + str(number_of_rec + 1)] = x['link']
    s = x['link'].split('?')[1]
    url = x['link'].split('?')[0]
    s1 = s.split('=')
    payload = {s1[0]: s1[1]}
    headers = {'user-agent': user_agent_string}
    page = requests.get(url, params=payload, headers=headers)
#    print(page.url)
#    print(page.status_code)
#    print(page.text)
    if page.status_code == requests.codes.ok:

        soup = BeautifulSoup(page.text, features="html.parser")
        all_td = soup.select('td')
        for i in range(len(all_td)):
            if all_td[i].getText() == 'Наименование закупки':
                print('Тема отбора: ' + all_td[i+1].getText().strip(' \n'))
                ws['D' + str(number_of_rec + 1)] = all_td[i+1].getText().strip(' \n')
            if all_td[i].getText() == 'Наименование организации':
                print('Организатор: ' + all_td[i+1].getText().strip(' \n'))
                ws['E' + str(number_of_rec + 1)] = all_td[i+1].getText().strip(' \n')
            if all_td[i].getText() == 'Регион':
                print('Регион: ' + all_td[i+1].getText(), end='  ')
                ws['F' + str(number_of_rec + 1)] = all_td[i+1].getText().strip(' \n')
            if all_td[i].getText() == 'Дата и время начала срока подачи заявок на участие в предварительном отборе':
                print('Начало подачи : ' + all_td[i+1].getText(), end='  ')
                ws['G' + str(number_of_rec + 1)] = all_td[i+1].getText().strip(' \n')
            if all_td[i].getText() == 'Дата и время окончания срока подачи заявок на участие в предварительном отборе':
                print('Конец подачи : ' + all_td[i+1].getText())
                ws['H' + str(number_of_rec + 1)] = all_td[i+1].getText().strip(' \n')

print('--------------------------------------------')
for col in ws['A2:H' + str(number_of_rec + 1)]:
    for cell in col:
        cell.style = 'styleraws'

wb.save('otbor615.xlsx')